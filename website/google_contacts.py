'''
Author: Naor Or-Zion
State: Israel
From: Basmach - Mador Liba
Date released: 15/05/2023

Brief: The code adds contacts to google contacts.
       The code is using Google People API and supports OAuth 2.0 only, there is no need for API key!
       Link to API integration overview: https://developers.google.com/people/api/rest/v1/people/
       Please read the README.txt for further information.
'''


import re
import os
import requests
import openpyxl
from flask import flash
from pathlib import Path
from zipfile import BadZipfile
from typing import Tuple, List, Union
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow


# Consts
# Local Paths
RESOURCES_DIR = Path("website/resources")
CLIENT_FILE = RESOURCES_DIR / "client_secret.json"
TOKEN_FILE = RESOURCES_DIR / "token.json"

# URLs
OAUTH_SCOPE = ['https://www.googleapis.com/auth/contacts']
CREATE_BATCH_CONTACTS_URL = "https://people.googleapis.com/v1/people:batchCreateContacts"
DELETE_BATCH_CONTACTS_URL = "https://people.googleapis.com/v1/people:batchDeleteContacts"
SEARCH_CONTACT_URL = "https://people.googleapis.com/v1/people:searchContacts"
LIST_CONTACTS_URL = "https://people.googleapis.com/v1/people/me/connections"

# TEXT
FINISH_TEXT = "All users have been added successfully"
PHONE_NUMBER_REGEX = r"^(?:(?:(\+?972|\(\+?972\)|\+?\(972\))(?:\s|\.|-)?([1-9]\d?))|(0[23489]{1})|(0[57]{1}[0-9]))(?:\s|\.|-)?([^0\D]{1}\d{2}(?:\s|\.|-)?\d{4})$"

# Network - Must use this proxy with any request
PROXY = {'http':'http://10.0.2.69:80'}
PROXY_IP = "10.0.2.69"
LOCAL_PORT_NUMBER = 62223
HTTP_OK_STATUS = 200

# Limitations
CONTACTS_CHUNK_LIMIT = 200


def get_credentials() -> Union[Credentials, str, bool]:
    """
    This function gets the user's credentials from google with the help of the OAuth 2.0 Scope.
    OAuth 2.0 holds an access token. Remember - The access token replaces the API key.

    If user's credentials exists, use the existed credentials.
    Eventually this function will return the credentials(aspecially the access token) which 
    will grant the user a communication with the API servers.

    :return credentials
    """
    creds = None

    if not os.path.isfile(CLIENT_FILE):
        flash(f"קובץ ה-client_secret.json לא נמצא!", "danger")
        return False
    
    # Checks if the creds file is exists in the same directory.
    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), OAUTH_SCOPE)

    # If the creds file is valid, return it.
    if creds and creds.valid:
        return creds
        
    # If the creds file is expired then refresh it - Request again    
    if creds and creds.expired and creds.refresh_token:
        try: 
            creds.refresh(Request())
            return creds
        except:
            "Proxy Error"

    # Use a Client File in the OAuth flow to acquire an access token
    # associated with your project on behalf of a user's account.
    try:
        flow = InstalledAppFlow.from_client_secrets_file(client_secrets_file=str(CLIENT_FILE), scopes=OAUTH_SCOPE)
    except OSError as e:
        flash(f"קובץ ה-client_secret.json לא נמצא!", "danger")
        return False
      
    try: 
        creds = flow.run_local_server(port=LOCAL_PORT_NUMBER)
        TOKEN_FILE.write_text(creds.to_json())
    except requests.exceptions.ConnectionError:
        return "Proxy Error"

    return creds   


def add_contacts(session: requests.Session, contacts_to_create: dict, contact_duplicates_to_delete: dict, headers: dict) -> bool:
    """
    This function creates a contact by a given name and a phone number.
    It first sets up the POST request parameters, then adds a contact.

    :session: The request session.
    :param contacts: A dictonary that contains all the contacts.
    :param headers: Neccesery headers for the POST request.
    """
    response_delete_batch_contacts = session.post(DELETE_BATCH_CONTACTS_URL, headers=headers, json=contact_duplicates_to_delete, proxies=PROXY)
    if response_delete_batch_contacts.status_code != HTTP_OK_STATUS:
        flash(f"אנשי הקשר נוספו בהצלחה ללא מחיקה של אנשי קשר אחרים", "success")

    response_create_batch_contacts = session.post(CREATE_BATCH_CONTACTS_URL, headers=headers, json=contacts_to_create, proxies=PROXY)
    if response_create_batch_contacts.status_code != HTTP_OK_STATUS:
        flash(f"Error creating contacts: {response_create_batch_contacts.text}", "danger")
        is_error = True
        return is_error
    
    return False


def add_contacts_from_excel(file_path: Path, key_words: str) -> Tuple[List[str], List[str], List[str], bool]:
    """
    This function gets a file path to the excel file containing the data of the contacts.
    The data in the excel file MUST be 2 columns only while the first column is "Full name" and the second column is "Phone number".
    The file must be in a left-to-right alignment format.
    The data can be as many rows as you want.
    Example of how it should look like in excel:
    +--------------------+--------------------+--------------------+
    |      Full name     |     Phone Number   |   Current Course   |
    +--------------------+--------------------+--------------------+
    |    Lionel Messi    |      0501010101    |        Python      |
    +--------------------+--------------------+--------------------+
    |    Ryan Rynolds    |      0506942069    |         SQL        |
    +--------------------+--------------------+--------------------+

    :param file_path: The path to the excel file containing the contacts.
    """
    is_error = False
    key_words = '' if key_words is None else str(key_words)

    # Lists of names, phone numbers and current course
    names = []
    phone_numbers = []
    current_courses = []
    contacts = { "contacts": [] }

    # Lists of existed names, phone numbers and resource names
    names_existed = []
    phone_numbers_existed = []
    resourceName_all_contacts = []
    resourceName_duplicate_contacts = { "resourceNames": [] }

    # Create a session with the requests library.
    google_session = requests.Session()

    # Create a contacts window instance and get the credentials to login and eventually a1dd the contacts to that user.
    credentials = get_credentials()

    # Check if there is a proxy error, if so, notify the client.
    if credentials == "Proxy Error":
        flash(f"צריך לשנות את הפרוקסי לכתובת: {PROXY_IP}", "danger")
        is_error = True
        return names, phone_numbers, current_courses, is_error
    #Check if there is any error, if so, notify the client.
    elif credentials is False:
        is_error = True
        return names, phone_numbers, current_courses, is_error

    # Set up the POST request parameters
    headers = {
        # 'credentials.token' is access token 
        'Authorization': f'Bearer {credentials.token}',
        'Accept': 'application/json',
        'Content-Type': 'application/json'
    }

    params_contacts_list = {
        "resourceName": "people/me",
        'personFields': 'names,phoneNumbers',
        'pageSize': 200
    }

    while True: 
        # Get the data of all existed contacts in json format.
        response_contacts_list = google_session.get(LIST_CONTACTS_URL, headers=headers, params=params_contacts_list)
        if response_contacts_list.status_code != HTTP_OK_STATUS:
                flash(f"Error listing contacts: {response_contacts_list.text}", "danger")
                is_error = True
                return names, phone_numbers, current_courses, is_error

        
        # Create 3 lists of contacts thats already exists: Name, phone number and resource name.
        if 'connections' in response_contacts_list.json():
            for contact in response_contacts_list.json()['connections']:
                names_dict = contact.get('names', [])
                if names_dict:
                    names_existed.append(names_dict[0]['displayName'])
                else:
                    names_existed.append("No name")

                phone_numbers_dict = contact.get('phoneNumbers', [])
                if phone_numbers_dict:
                    value = phone_numbers_dict[0]['value']
                    # Clean white spaces
                    value = value.replace(" ", "")
                    # Replace +972 with 0
                    value = value.replace("+972", "0")
                    # Transform the number from: "+972 50-123-4567" to "0501234567"
                    value = value.replace("-", "")
                    phone_numbers_existed.append(value)
                else:
                    phone_numbers_existed.append("No phone number")

                resourceName_all_contacts.append(contact['resourceName'])

        # If there are no more pages, break out of the loop
        if 'nextPageToken' not in response_contacts_list.json():
            break

        # Set the page token to the next page token to retrieve the next page of results
        params_contacts_list['pageToken'] = response_contacts_list.json()['nextPageToken']


    # Load the Excel file and assign the number of rows and the number of columns to 2 appropriate variables.
    try:
        workbook = openpyxl.load_workbook(file_path)
    except BadZipfile:
        flash("קובץ האקסל שבחרת לא תקין, צריך קובץ xlsx", "danger")
        is_error = True
        return names, phone_numbers, current_courses, is_error

    worksheet = workbook.active
    row_count = worksheet.max_row
    column_count = worksheet.max_column

    # Checks if number of rows is less than 2, if so, show an error message and quit the function
    if row_count < 2:
        flash("צריך לבחור קובץ אקסל שיש לו יותר משורה אחת", "danger")
        is_error = True
        return names, phone_numbers, current_courses, is_error
    
    # Checks if number of columns is exactly 3, if so, show an error message and quit the function
    if column_count != 3:
        flash("צריך לבחור קובץ אקסל עם 3 עמודות בדיוק", "danger")
        is_error = True
        return names, phone_numbers, current_courses, is_error

    # Extract the name, phone number and the current course for each row in the sheet, then:
    # Create lists for name, phone numbers and current course.
    # Create a json style dictionary designed to contain all the possible names, phone numbers and current courses - 
    # this dictionary will be sent in the request headers to create contacts. 
    for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
        name, phone_number, current_course = row

        # When a name cell is empty raise an error
        if not name:
            flash("יש שם ריק בטופס! אנשי הקשר נוספו עד אותו שם", "danger")
            is_error = True
            return names, phone_numbers, current_courses, is_error
        
        # When a phone_number cell is empty raise an error
        if not phone_number:
            flash("יש תא ללא טלפון בטופס! אנשי הקשר נוספו עד אותו שם", "danger")
            is_error = True
            return names, phone_numbers, current_courses, is_error
        
        # When a current_course cell is empty raise an error
        if not current_course:
            flash("יש תא ללא שם קורס בטופס! אנשי הקשר נוספו עד אותו שם", "danger")
            is_error = True
            return names, phone_numbers, current_courses, is_error
        
        name = str(name)
        phone_number = str(phone_number)
        current_course = str(current_course)

        is_phone_number_valid = re.match(PHONE_NUMBER_REGEX, phone_number)

        if is_phone_number_valid:
            # Clean white spaces
            phone_number.replace(" ", "")
            # Replace + with none
            phone_number.replace("+", "")
            # Replace ( with none
            phone_number.replace("(", "")
            # Replace ) with none
            phone_number.replace(")", "")
            # Replace 972 with 0
            phone_number.replace("972", "0")
            # Transform the number from: "+972 50-123-4567" to "0501234567"
            phone_number = phone_number.replace("-", "")
            # If phone number length is less than 10 digits, add 0 at the start of it
            phone_number = "0" + phone_number if len(phone_number) < 10 else phone_number
        else:
            flash("קיים מספר טלפון לא תקין בקובץ!", "danger")
            is_error = True
            return names, phone_numbers, current_courses, is_error

        # If a contact is already exists add him to the delete-duplicates list and move to next iteration.
        if phone_number in phone_numbers_existed:
            index = phone_numbers_existed.index(phone_number)
            resourceName_duplicate_contacts["resourceNames"].append(resourceName_all_contacts[index])

        names.append(name + " " + key_words)
        phone_numbers.append(phone_number)
        current_courses.append(current_course)

        contacts["contacts"].append(
            {
                "contactPerson": 
                {
                    "names": [
                        {
                            "givenName": name + " " + key_words
                        }
                    ],
                    "phoneNumbers": [
                        {
                            "value": phone_number,
                            "type": "mobile"
                        }
                    ],
                    "organizations": [
                        {
                            "name": current_course
                        }
                    ]           
                }
            }
        )
        
        
        # There is a CONTACTS_CHUNK_LIMIT contacts limit when requesting to create new contacts.
        # The contacts are divided into chunks of CONTACTS_CHUNK_LIMIT to not exceed the limit.
        # The 'index + 1' is there because the index starts from 0.
        if not ((index + 1) % CONTACTS_CHUNK_LIMIT):
            is_error = add_contacts(google_session, contacts, resourceName_duplicate_contacts, headers)
            contacts["contacts"] = []
            if is_error:
                return names, phone_numbers, current_courses, is_error
    
    # Send the remaining contacts which are left when not divisible by CONTACTS_CHUNK_LIMIT.
    if contacts["contacts"]:
        is_error = add_contacts(google_session, contacts, resourceName_duplicate_contacts, headers)
        if is_error:
            return names, phone_numbers, current_courses, is_error
    

    return names, phone_numbers, current_courses, is_error
